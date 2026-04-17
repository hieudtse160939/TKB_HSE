import pandas as pd

import openpyxl

import streamlit as st

import io

import re



# =====================================================================

# 1. BỘ TỪ ĐIỂN TỰ ĐỘNG CỦA TRƯỜNG HOA SEN

# =====================================================================

# =====================================================================
# BỘ TỪ ĐIỂN TỰ ĐỘNG ĐÃ ĐƯỢC MAPPING TỪ FILE TUẦN 1 & DSGVBM
# =====================================================================
TU_DIEN = {
    # --- TỔ VĂN ---
    ("Việt", "V"): "Thầy Lê Công Quốc Việt (Văn)",
    ("Việt", "CĐ Văn"): "Thầy Lê Công Quốc Việt (Văn)",
    ("Nhung", "V"): "Cô Trần Thị Tuyết Nhung (Văn)",
    ("Nhung", "CĐ Văn"): "Cô Trần Thị Tuyết Nhung (Văn)",
    ("Tâm", "V"): "Thầy Ngô Đình Tâm (Văn)",
    ("Tâm", "CĐ Văn"): "Thầy Ngô Đình Tâm (Văn)",
    ("Bình", "V"): "Cô Lê Thị Bình (Văn)",
    ("Bình", "CĐ Văn"): "Cô Lê Thị Bình (Văn)",
    ("Hạnh", "V"): "Cô Lê Ngọc Bích Hạnh (Văn)",
    ("Hạnh", "CĐ Văn"): "Cô Lê Ngọc Bích Hạnh (Văn)",
    ("Nhàn", "V"): "Cô Lưu Thị Nhàn (Văn)",
    ("Nhàn", "CĐ Văn"): "Cô Lưu Thị Nhàn (Văn)",
    ("Thoa", "V"): "Cô Nguyễn Thị Thoa (Văn)",
    ("Thoa", "CĐ Văn"): "Cô Nguyễn Thị Thoa (Văn)",
    ("Oanh", "V"): "Cô Võ Thị Oanh (Văn)",
    ("Oanh", "CĐ Văn"): "Cô Võ Thị Oanh (Văn)",
    ("Như", "V"): "Cô Võ Nguyễn Huỳnh Như (Văn)",
    ("Như", "CĐ Văn"): "Cô Võ Nguyễn Huỳnh Như (Văn)",
    ("Diệp", "V"): "Cô Phan Thị Ngọc Diệp (Văn)",
    ("Diệp", "CĐ Văn"): "Cô Phan Thị Ngọc Diệp (Văn)",
    ("Nam", "V"): "Thầy Nguyễn Văn Nam (Văn)",

    # --- TỔ TOÁN ---
    ("Nghĩa", "T"): "Thầy Nghĩa (Toán)",
    ("Nghĩa", "CĐ Toán"): "Thầy Nghĩa (Toán)",
    ("Như", "T"): "Cô Nguyễn Trương Quỳnh Như (Toán)",
    ("Như", "CĐ Toán"): "Cô Nguyễn Trương Quỳnh Như (Toán)",
    ("Nhi", "T"): "Cô Lê Thị Yến Nhi (Toán)",
    ("Nhi", "CĐ Toán"): "Cô Lê Thị Yến Nhi (Toán)",
    ("Sang", "T"): "Thầy Huỳnh Phước Sang (Toán)",
    ("Sang", "Toán"): "Thầy Huỳnh Phước Sang (Toán)",
    ("Sanh", "T"): "Thầy Đặng Văn Sanh (Toán)",
    ("Sanh", "CĐ Toán"): "Thầy Đặng Văn Sanh (Toán)",
    ("Vũ", "T"): "Thầy Hồ Thế Vũ (Toán)",
    ("Vũ", "CĐ Toán"): "Thầy Hồ Thế Vũ (Toán)",
    ("Thọ", "T"): "Thầy Phan Minh Thọ (Toán)",
    ("Thọ", "CĐ Toán"): "Thầy Phan Minh Thọ (Toán)",
    ("Thuận", "T"): "Thầy Lư Quang Thuận (Toán)",
    ("Thuận", "CĐ Toán"): "Thầy Lư Quang Thuận (Toán)",
    ("Thương", "T"): "Cô Trần Thị Thanh Thương (Toán)",
    ("Thương", "CĐ Toán"): "Cô Trần Thị Thanh Thương (Toán)",
    ("Trinh", "T"): "Cô Nguyễn Thị Bích Trinh (Toán)",
    ("Ánh", "T"): "Cô Lê Thị Ngọc Ánh (Toán)",
    ("Ánh", "CĐ Toán"): "Cô Lê Thị Ngọc Ánh (Toán)",

    # --- TỔ ANH VĂN ---
    ("Vân", "A"): "Cô Nguyễn Thảo Vân (Anh)",
    ("Vân", "IELTS/A"): "Cô Nguyễn Thảo Vân (Anh)",
    ("Lài", "A"): "Cô Nguyễn Thị Lài (Anh)",
    ("Lài", "IELTS/A"): "Cô Nguyễn Thị Lài (Anh)",
    ("Nhi", "A"): "Cô Trần Trúc Nhi (Anh)",
    ("Nhung", "A"): "Cô Nguyễn Thị Phương Nhung (Anh)",
    ("Biền", "A"): "Cô Đinh Thị Biền (Anh)",
    ("Giang", "A"): "Cô Bùi Thị Giang (Anh)",
    ("Loan", "A"): "Cô Chu Thị Hồng Loan (Anh)",
    ("Thúy", "A"): "Cô Ngô Hà Thanh Thúy (Anh)",
    ("Thư", "A"): "Cô Phan Thị Minh Thư (Anh)",
    ("Thu", "A"): "Cô Thái Ngọc Thu (Anh)",
    ("Đào", "A"): "Cô Lê Hồng Đào (Anh)",
    ("Đông", "A"): "Cô Trần Thị Đông (Anh)",

    # --- TỔ LÝ - HÓA - SINH - CN ---
    ("Vân", "L"): "Cô Nguyễn Thị Vân (Lý)",
    ("Vân", "CĐ Lý"): "Cô Nguyễn Thị Vân (Lý)",
    ("Ngọc", "L"): "Cô Phạm Thị Cẩm Ngọc (Lý)",
    ("Ngọc", "CĐ Lý"): "Cô Phạm Thị Cẩm Ngọc (Lý)",
    ("Ngân", "L"): "Cô Hà Thị Kim Ngân (Lý)",
    ("Ngân", "CĐ Lý"): "Cô Hà Thị Kim Ngân (Lý)",
    ("Quỳnh", "L"): "Cô Mai Thị Ngọc Quỳnh (Lý)",
    ("Đức", "L"): "Thầy Nguyễn Vĩnh Đức (Lý)",
    ("Đức", "CĐ Lý"): "Thầy Nguyễn Vĩnh Đức (Lý)",
    ("Chi", "H"): "Cô Trần Thị Quí Chi (Hóa)",
    ("Phượng", "H"): "Cô Vũ Thị Phượng (Hóa)",
    ("Quỳnh", "H"): "Cô Nguyễn Thị Như Quỳnh (Hóa)",
    ("Ân", "H"): "Thầy Nguyễn Hoàng Thiên Ân (Hóa)",
    ("Thi", "Si"): "Thầy Trương Đình Thi (Sinh)",
    ("Trinh", "Si"): "Cô Huỳnh Thị Tuyết Trinh (Sinh)",
    ("Hải", "Si"): "Cô Bùi Thị Minh Hải (Sinh)",

    # --- TỔ SỬ - ĐỊA - KTPL ---
    ("Bảo", "Su"): "Thầy Vương Quốc Bảo (Sử)",
    ("Phương", "Su"): "Cô Tống Thị Phương (Sử)",
    ("Trung", "Su"): "Thầy Lý Văn Trung (Sử)",
    ("Thủy", "Su"): "Cô Đỗ Thị Thu Thủy (Sử)",
    ("Hương", "Su"): "Cô Nguyễn Thị Hương (Sử)",
    ("Linh", "Su"): "Cô Nguyễn Thị Linh (Sử)",
    ("Vy", "Đ"): "Cô Phan Thị Thảo Vy (Địa)",
    ("Hương", "Đ"): "Cô Võ Thị Hương (Địa)",
    ("Hiền", "Đ"): "Cô Đặng Thị Hiền (Địa)",
    ("Minh", "Đ"): "Cô Mai Thị Ngọc Minh (Địa)",
    ("Xuân", "KTPL"): "Cô Xuân (KTPL)",
    ("Vinh", "KTPL"): "Cô Vinh (KTPL)",
    ("Anh", "KTPL"): "Cô Lan Anh (KTPL)",
    ("Ngọc", "KTPL"): "Thầy Ngọc (KTPL)",

    # --- CÁC TRƯỜNG HỢP ĐẢO NGƯỢC (Tên - Môn) ---
    # (Dùng để fix lỗi TKB ghi ngược "Tên giáo viên - Tên môn")
    ("V", "Việt"): "Thầy Lê Công Quốc Việt (Văn)",
    ("V", "Nhung"): "Cô Trần Thị Tuyết Nhung (Văn)",
    ("V", "Tâm"): "Thầy Ngô Đình Tâm (Văn)",
    ("T", "Nghĩa"): "Thầy Nghĩa (Toán)",
    ("L", "Vân"): "Cô Nguyễn Thị Vân (Lý)",
    ("L", "Ngọc"): "Cô Phạm Thị Cẩm Ngọc (Lý)",
    ("A", "Vân"): "Cô Nguyễn Thảo Vân (Anh)",
}



def get_standard_name(ten_ky_hieu, mon_hoc):

    if mon_hoc == 'Chủ nhiệm':

        return f"{ten_ky_hieu} (GVCN)"

    if (ten_ky_hieu, mon_hoc) in TU_DIEN:

        return TU_DIEN[(ten_ky_hieu, mon_hoc)]

    return ten_ky_hieu



def phan_loai_khoi(ten_lop):

    """Hàm tự động phân loại Khối THCS và THPT"""

    ten_lop = str(ten_lop).strip()

    if ten_lop.startswith(('10', '11', '12')):

        return 'THPT'

    elif ten_lop.startswith(('6', '7', '8', '9')):

        return 'THCS'

    else:

        return 'Khác'



# =====================================================================

# 2. HÀM XỬ LÝ CHÍNH TRÊN MEMORY

# =====================================================================

def process_tkb_data(uploaded_file):

    wb = openpyxl.load_workbook(uploaded_file)

    

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        merged_ranges = list(ws.merged_cells.ranges)

        

        for merged_range in merged_ranges:

            min_col, min_row, max_col, max_row = merged_range.bounds

            top_left_cell_value = ws.cell(row=min_row, column=min_col).value

            ws.unmerge_cells(str(merged_range))

            for row in range(min_row, max_row + 1):

                for col in range(min_col, max_col + 1):

                    ws.cell(row=row, column=col).value = top_left_cell_value

                    

        max_row_current = ws.max_row

        if max_row_current > 66:

            ws.delete_rows(67, max_row_current - 66)

            

    virtual_workbook = io.BytesIO()

    wb.save(virtual_workbook)

    virtual_workbook.seek(0)

    

    df = pd.read_excel(virtual_workbook) 

    

    class_row_idx = 3

    gvcn_row_idx = 4

    classes = {} 

    gvcn = {}    

    

    class_row = df.iloc[class_row_idx]

    gvcn_row = df.iloc[gvcn_row_idx]

    

    for col_idx, val in enumerate(class_row):

        # Regex CHỈ bắt tên lớp bắt đầu bằng số từ 6 đến 12 (6, 7, 8, 9, 10, 11, 12)

        if pd.notna(val) and isinstance(val, str) and re.match(r'^(1[0-2]|[6-9])', val.strip()): 

            class_name = val.strip()

            classes[col_idx] = class_name

            gv_val = str(gvcn_row.iloc[col_idx]).strip()

            if '-' in gv_val:

                gvcn[class_name] = gv_val.split('-')[0].strip()

            else:

                gvcn[class_name] = gv_val



    records = []

    for row_idx in range(5, len(df)):

        row = df.iloc[row_idx]

        for col_idx, class_name in classes.items():

            cell_val = str(row.iloc[col_idx]).strip()

            if cell_val == 'nan' or cell_val == '' or cell_val in ['CHÀO CỜ', 'SINH HOẠT ĐẦU GIỜ', 'THỂ DỤC THỂ THAO']:

                continue

            

            khoi = phan_loai_khoi(class_name)



            if cell_val.lower() == 'chủ nhiệm':

                ten_goc = gvcn.get(class_name, 'Unknown')

                records.append({

                    'Khối': khoi,

                    'Giáo viên': get_standard_name(ten_goc, 'Chủ nhiệm'),

                    'Lớp': class_name, 

                    'Môn': 'Chủ nhiệm'

                })

            elif '-' in cell_val:

                parts = cell_val.split('-')

                if len(parts) >= 2:

                    mon_hoc = "-".join(parts[:-1]).strip() 

                    ten_goc = parts[-1].strip()        

                    records.append({

                        'Khối': khoi,

                        'Giáo viên': get_standard_name(ten_goc, mon_hoc),

                        'Lớp': class_name, 

                        'Môn': mon_hoc

                    })

            else:

                records.append({

                    'Khối': khoi,

                    'Giáo viên': 'Chung (Không ghi tên)',

                    'Lớp': class_name, 

                    'Môn': cell_val

                })



    df_records = pd.DataFrame(records)

    if df_records.empty:

        return None

        

    df_summary = df_records.groupby(['Khối', 'Giáo viên', 'Lớp', 'Môn']).size().reset_index(name='Số tiết')

    df_summary = df_summary.sort_values(by=['Khối', 'Giáo viên', 'Lớp', 'Môn'])

    

    return df_summary



# =====================================================================

# 3. GIAO DIỆN STREAMLIT VỚI FILTER VÀ EXPORT

# =====================================================================

st.set_page_config(page_title="Thống kê TKB", page_icon="📊", layout="wide")



st.title("📊 Công cụ Xử lý & Thống kê TKB")



# --- KHỞI TẠO SESSION STATE ---

if 'df_data' not in st.session_state:

    st.session_state.df_data = None



# --- KHU VỰC UPLOAD ---

uploaded_file = st.file_uploader("Kéo thả hoặc chọn file TKB (.xlsx) vào đây", type=["xlsx"])



if uploaded_file is not None:

    if st.button("Bắt đầu phân tích dữ liệu", type="primary"):

        with st.spinner('Đang xử lý dữ liệu...'):

            df_ket_qua = process_tkb_data(uploaded_file)

            

            if df_ket_qua is not None:

                st.session_state.df_data = df_ket_qua

                st.success("Đã phân tích xong! Bạn có thể dùng bộ lọc bên dưới.")

            else:

                st.error("Không tìm thấy dữ liệu hợp lệ trong file Excel.")



# --- KHU VỰC BỘ LỌC & HIỂN THỊ ---

if st.session_state.df_data is not None:

    df = st.session_state.df_data

    

    st.divider()

    st.subheader("🔍 Bộ lọc dữ liệu")

    

    col1, col2, col3 = st.columns(3)

    

    with col1:

        danh_sach_khoi = sorted(df['Khối'].unique().tolist())

        khoi_chon = st.multiselect("📚 Chọn Khối (Để trống để xem tất cả):", danh_sach_khoi)



    with col2:

        danh_sach_gv = sorted(df['Giáo viên'].unique().tolist())

        gv_chon = st.multiselect("👩‍🏫 Chọn Giáo viên:", danh_sach_gv)

        

    with col3:

        danh_sach_lop = sorted(df['Lớp'].unique().tolist())

        lop_chon = st.multiselect("🏫 Chọn Lớp:", danh_sach_lop)
